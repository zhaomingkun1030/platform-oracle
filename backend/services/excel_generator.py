"""
Excel Report Generator - Excel 报告生成服务
"""

import os
from datetime import datetime
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Excel 报告生成器"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(
        self,
        analysis_result: Dict[str, Any],
        predictor_result: Dict[str, Any],
        output_path: str,
        url: str,
        analyst: str
    ):
        """
        生成 Excel 报告
        
        Args:
            analysis_result: Analysis Agent 分析结果
            predictor_result: Predictor Agent 预测结果
            output_path: 输出文件路径
            url: 分析的 URL
            analyst: 分析人员
        """
        wb = Workbook()
        
        # Sheet 1: 功能点
        self._create_features_sheet(wb, analysis_result)
        
        # Sheet 2: 策略解释
        self._create_strategy_sheet(wb, analysis_result)
        
        # Sheet 3: Roadmap 预测
        self._create_roadmap_sheet(wb, predictor_result)
        
        # Sheet 4: 执行元数据
        self._create_metadata_sheet(wb, url, analyst, analysis_result, predictor_result)
        
        # 保存
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)
    
    def _create_features_sheet(self, wb: Workbook, analysis_result: Dict[str, Any]):
        """创建功能点 Sheet"""
        ws = wb.active
        ws.title = "功能点"
        
        # 标题
        headers = ["功能名称", "功能描述", "开源组件", "使用场景"]
        ws.append(headers)
        
        # 格式化标题行
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        # 添加数据
        features = analysis_result.get("features", [])
        for feature in features:
            ws.append([
                feature.get("name", ""),
                feature.get("description", ""),
                ", ".join(feature.get("open_source_components", [])),
                feature.get("use_case", "")
            ])
        
        # 调整列宽
        self._adjust_column_width(ws)
    
    def _create_strategy_sheet(self, wb: Workbook, analysis_result: Dict[str, Any]):
        """创建策略解释 Sheet"""
        ws = wb.create_sheet("策略解释")
        
        headers = ["功能名称", "竞品策略解释"]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
        
        features = analysis_result.get("features", [])
        for feature in features:
            ws.append([
                feature.get("name", ""),
                feature.get("competitor_strategy", "")
            ])
        
        # 添加整体总结
        summary = analysis_result.get("summary", "")
        if summary:
            ws.append([])
            ws.append(["整体分析总结"])
            ws.append([summary])
        
        self._adjust_column_width(ws)
    
    def _create_roadmap_sheet(self, wb: Workbook, predictor_result: Dict[str, Any]):
        """创建 Roadmap Sheet"""
        ws = wb.create_sheet("Roadmap 预测")
        
        # 技术趋势
        ws.append(["技术趋势预测"])
        headers = ["趋势", "描述", "可信度"]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        trends = predictor_result.get("technology_trends", [])
        for trend in trends:
            ws.append([
                trend.get("trend", ""),
                trend.get("description", ""),
                trend.get("confidence", "")
            ])
        
        ws.append([])
        
        # 能力方向
        ws.append(["能力方向预测"])
        headers = ["方向", "描述", "时间线"]
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row - 1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        directions = predictor_result.get("capability_directions", [])
        for direction in directions:
            ws.append([
                direction.get("direction", ""),
                direction.get("description", ""),
                direction.get("timeline", "")
            ])
        
        ws.append([])
        
        # 战略预测
        ws.append(["战略预测"])
        predictions = predictor_result.get("strategic_predictions", [])
        for pred in predictions:
            ws.append([pred])
        
        self._adjust_column_width(ws)
    
    def _create_metadata_sheet(
        self,
        wb: Workbook,
        url: str,
        analyst: str,
        analysis_result: Dict[str, Any],
        predictor_result: Dict[str, Any]
    ):
        """创建元数据 Sheet"""
        ws = wb.create_sheet("执行元数据")
        
        ws.append(["Platform Oracle - 分析报告"])
        ws.append([])
        
        metadata = [
            ("分析 URL", url),
            ("分析人员", analyst),
            ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("来源", analysis_result.get("source_url", "")),
            ("数据类型", "视频" if analysis_result.get("is_video") else "文档"),
            ("功能点数量", len(analysis_result.get("features", []))),
            ("关键洞察", ", ".join(analysis_result.get("key_insights", []))),
        ]
        
        for key, value in metadata:
            ws.append([key, value])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
    
    def _adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
